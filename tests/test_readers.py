import pytest
from openpyxl import Workbook

from deportation_foia_normalizer.readers import read_input


class TestReadCSV:
    def test_csv_basic(self, tmp_path):
        csv_file = tmp_path / "test.csv"
        csv_file.write_text("name,age,city\nAlice,30,NYC\nBob,25,LA\n")

        headers, rows = read_input(csv_file)

        assert headers == ["name", "age", "city"]
        assert rows == [
            {"name": "Alice", "age": "30", "city": "NYC"},
            {"name": "Bob", "age": "25", "city": "LA"},
        ]

    def test_csv_none_cells_become_empty_strings(self, tmp_path):
        csv_file = tmp_path / "test.csv"
        csv_file.write_text("name,age\nAlice,30\nBob,\n")

        headers, rows = read_input(csv_file)

        assert rows[1]["age"] == ""

    def test_csv_extra_fields_ignored(self, tmp_path):
        csv_file = tmp_path / "test.csv"
        csv_file.write_text("name,age\nAlice,30,extra\n")

        headers, rows = read_input(csv_file)

        assert rows[0] == {"name": "Alice", "age": "30"}
        assert None not in rows[0]


class TestReadXLSX:
    def test_xlsx_basic(self, tmp_path):
        xlsx_file = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["name", "age", "city"])
        ws.append(["Alice", 30, "NYC"])
        ws.append(["Bob", 25, "LA"])
        wb.save(xlsx_file)

        headers, rows = read_input(xlsx_file)

        assert headers == ["name", "age", "city"]
        assert rows == [
            {"name": "Alice", "age": "30", "city": "NYC"},
            {"name": "Bob", "age": "25", "city": "LA"},
        ]

    def test_xlsx_none_cells_become_empty_strings(self, tmp_path):
        xlsx_file = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["name", "age"])
        ws.append(["Alice", 30])
        ws.append(["Bob", None])
        wb.save(xlsx_file)

        headers, rows = read_input(xlsx_file)

        assert rows[1]["age"] == ""

    def test_xlsx_empty_file(self, tmp_path):
        xlsx_file = tmp_path / "empty.xlsx"
        wb = Workbook()
        wb.save(xlsx_file)

        with pytest.raises(ValueError, match="XLSX file is empty"):
            read_input(xlsx_file)


class TestCSVandXLSXEquivalence:
    def test_csv_and_xlsx_identical_content(self, tmp_path):
        csv_file = tmp_path / "test.csv"
        csv_file.write_text("name,age,city\nAlice,30,NYC\nBob,25,LA\n")

        xlsx_file = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["name", "age", "city"])
        ws.append(["Alice", "30", "NYC"])
        ws.append(["Bob", "25", "LA"])
        wb.save(xlsx_file)

        csv_headers, csv_rows = read_input(csv_file)
        xlsx_headers, xlsx_rows = read_input(xlsx_file)

        assert csv_headers == xlsx_headers
        assert csv_rows == xlsx_rows


class TestErrorHandling:
    def test_missing_file_raises_filenotfounderror(self):
        with pytest.raises(FileNotFoundError):
            read_input("/nonexistent/file.csv")

    def test_unsupported_extension_raises_valueerror(self, tmp_path):
        unsupported_file = tmp_path / "test.json"
        unsupported_file.write_text("{}")

        with pytest.raises(ValueError, match="Unsupported file extension"):
            read_input(unsupported_file)

    def test_corrupted_csv_raises(self, tmp_path):
        csv_file = tmp_path / "test.csv"
        csv_file.write_text("name,age\nAlice,30")

        headers, rows = read_input(csv_file)
        assert headers == ["name", "age"]
        assert len(rows) == 1

    def test_corrupted_xlsx_raises(self, tmp_path):
        xlsx_file = tmp_path / "test.xlsx"
        xlsx_file.write_bytes(b"not a valid xlsx file")

        with pytest.raises(ValueError, match="Failed to parse XLSX"):
            read_input(xlsx_file)

    def test_csv_empty_file_has_no_headers(self, tmp_path):
        csv_file = tmp_path / "empty.csv"
        csv_file.write_text("")

        with pytest.raises(ValueError, match="CSV file is empty"):
            read_input(csv_file)


class TestValueNormalization:
    def test_all_values_converted_to_strings(self, tmp_path):
        xlsx_file = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["col1", "col2", "col3"])
        ws.append([123, 45.67, True])
        wb.save(xlsx_file)

        headers, rows = read_input(xlsx_file)

        row = rows[0]
        assert row["col1"] == "123"
        assert row["col2"] == "45.67"
        assert row["col3"] == "True"
        assert all(isinstance(v, str) for v in row.values())

    def test_csv_values_are_already_strings(self, tmp_path):
        csv_file = tmp_path / "test.csv"
        csv_file.write_text("col1,col2\n123,45.67\n")

        headers, rows = read_input(csv_file)

        row = rows[0]
        assert all(isinstance(v, str) for v in row.values())
